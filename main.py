import sys
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5 import uic
import pandas as pd
import matplotlib.pyplot as plt
from data_manager import DataManager
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from PyQt5.QtGui import *
from dateutil.relativedelta import relativedelta
from datetime import datetime
import matplotlib as mpl
import webbrowser
from openpyxl import load_workbook

form_class = uic.loadUiType("stock.ui")[0]

class WindowClass(QMainWindow, form_class):
    def __init__(self):
        super().__init__()
        self.initUI()
        self.setupUi(self)

        #gui상 메뉴bar생성
        menubar = self.menuBar()
        self.menubar.setStyleSheet("background-color: rgb(150,150,150); color: rgb(255,255,255); border: 1px solid #000")
        Filemenu = menubar.addMenu("버전업데이트")
        Filemenu1 = menubar.addMenu("종료")
        gitgubfile = QAction(QIcon('data_image.png'),'Github 연결', self)
        exit = QAction(QIcon('exit.png'), 'Exit', self)
        exit.setShortcut('Ctrl+Q')
        gitgubfile.triggered.connect(self.connect_git)
        exit.triggered.connect(qApp.quit)
        Filemenu.addAction(gitgubfile)
        Filemenu1.addAction(exit)

        # 기간 정보 제공
        wb = load_workbook('data_202111.xlsx')
        ws = wb.active
        A1 = ws["A2"].value
        A1_date = "{}-{}-{}".format(A1[:4], A1[4:6], A1[6:])
        x = ws.max_row
        A2 = ws.cell(row=x, column=1).value
        A2_date = "{}-{}-{}".format(A2[:4], A2[4:6], A2[6:])
        self.label_3.setObjectName("label")
        self.label_3.setText('검색가능 기간   ' + A2_date + '  ~  ' + A1_date)

        # 테이블위젯 스타일 변경
        self.tableWidget.horizontalHeader().setStyleSheet(
            "QHeaderView::section {background-color:#404040;color:#FFFFFF;}")
        self.tableWidget.setShowGrid(False)
        self.tableWidget.setStyleSheet('QTableView::item {border-bottom: 1px solid #d6d9dc;}')

        self.setWindowTitle("수급 주체별 Stock Back Testing")
        self.setWindowIcon(QIcon("icon1.png"))

        self.dfClose = pd.read_excel('data_202111.xlsx', sheet_name='종가')
        self.dfClose['일자'] = pd.to_datetime(self.dfClose['일자'].astype(str), format='%Y%m%d')

        self.dfkospi = pd.read_excel('data_202111.xlsx', sheet_name='kospi')
        self.dfkospi['일자'] = pd.to_datetime(self.dfkospi['일자'].astype(str), format='%Y%m%d')

        self.data_manager = DataManager(self.dfClose)
        self.result_arr = []
        self.fig = plt.figure(figsize=(20, 40))
        self.canvas = FigureCanvas(self.fig)
        self.verticalLayout.addWidget(self.canvas)

        # 시작/종료 날짜
        self.date_start.dateChanged.connect(self.dateStartFunction)
        self.date_end.dateChanged.connect(self.dateEndFunction)
        self.date_start.setDate(QDate(2021, 1, 1))
        self.date_end.setDate(QDate(2021, 11, 30))

        # 매수/매도 버튼
        self.radio_buy.clicked.connect(self.radioBuysellFunction)
        self.radio_sell.clicked.connect(self.radioBuysellFunction)

        # 검색 버튼
        self.btn_search.clicked.connect(self.btnsearchFunction)

        # 라디오 버튼
        self.radio_supplier_1.clicked.connect(self.radioSupplierFunction)
        self.radio_supplier_2.clicked.connect(self.radioSupplierFunction)
        self.radio_supplier_3.clicked.connect(self.radioSupplierFunction)

        # Top5 table
        self.tableWidget.clicked.connect(self.selectTableFunction)

        # 버튼1~11클릭
        self.pushButton.clicked.connect(self.btnClick1)
        self.pushButton_2.clicked.connect(self.btnClick2)
        self.pushButton_3.clicked.connect(self.btnClick3)
        self.pushButton_4.clicked.connect(self.btnClick4)
        self.pushButton_5.clicked.connect(self.btnClick5)
        self.pushButton_6.clicked.connect(self.btnClick6)
        self.pushButton_7.clicked.connect(self.btnClick7)
        self.pushButton_8.clicked.connect(self.btnClick8)
        self.pushButton_9.clicked.connect(self.btnClick9)
        self.pushButton_10.clicked.connect(self.btnClick10)
        self.pushButton_11.clicked.connect(self.save_graph)

        self.type = 0

        self.pushButton.hide()
        self.pushButton_2.hide()
        self.pushButton_3.hide()
        self.pushButton_4.hide()
        self.pushButton_5.hide()
        self.pushButton_6.hide()
        self.pushButton_7.hide()
        self.pushButton_8.hide()
        self.pushButton_9.hide()
        self.pushButton_10.hide()
        self.pushButton_11.hide()

    def initUI(self):
        self.lbl = QLabel(self)
        self.show()

    # 버튼 1번 클릭시 그래프 생성
    def btnClick1(self):
        self.draw_graph(1)
        self.pushButton.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                      "background-color: rgb(255, 130, 130);")
        if self.type == 2:
            self.pushButton_2.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                            "background-color: rgb(230, 230, 230);")
        elif self.type == 3:
            self.pushButton_3.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                            "background-color: rgb(230, 230, 230);")
        elif self.type == 4:
            self.pushButton_4.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                            "background-color: rgb(230, 230, 230);")
        elif self.type == 5:
            self.pushButton_5.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                            "background-color: rgb(230, 230, 230);")
        elif self.type == 6:
            self.pushButton_6.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                            "background-color: rgb(230, 230, 230);")
        elif self.type == 7:
            self.pushButton_7.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                            "background-color: rgb(230, 230, 230);")
        elif self.type == 8:
            self.pushButton_8.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                            "background-color: rgb(230, 230, 230);")
        elif self.type == 9:
            self.pushButton_9.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                            "background-color: rgb(230, 230, 230);")
        elif self.type == 10:
            self.pushButton_10.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                             "background-color: rgb(230, 230, 230);")
        else:
            pass
        self.type = 1

    # 버튼 2번 클릭시 그래프 생성
    def btnClick2(self):
        self.draw_graph(2)
        self.pushButton_2.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                        "background-color: rgb(255, 130, 130);")
        if self.type == 1:
            self.pushButton.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                          "background-color: rgb(230, 230, 230);")
        elif self.type == 3:
            self.pushButton_3.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                            "background-color: rgb(230, 230, 230);")
        elif self.type == 4:
            self.pushButton_4.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                            "background-color: rgb(230, 230, 230);")
        elif self.type == 5:
            self.pushButton_5.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                            "background-color: rgb(230, 230, 230);")
        elif self.type == 6:
            self.pushButton_6.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                            "background-color: rgb(230, 230, 230);")
        elif self.type == 7:
            self.pushButton_7.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                            "background-color: rgb(230, 230, 230);")
        elif self.type == 8:
            self.pushButton_8.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                            "background-color: rgb(230, 230, 230);")
        elif self.type == 9:
            self.pushButton_9.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                            "background-color: rgb(230, 230, 230);")
        elif self.type == 10:
            self.pushButton_10.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                             "background-color: rgb(230, 230, 230);")
        else:
            pass
        self.type = 2

    # 버튼 3번 클릭시 그래프 생성
    def btnClick3(self):
        self.draw_graph(3)
        self.pushButton_3.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                        "background-color: rgb(255, 130, 130);")
        if self.type == 1:
            self.pushButton.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                          "background-color: rgb(230, 230, 230);")
        elif self.type == 2:
            self.pushButton_2.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                            "background-color: rgb(230, 230, 230);")
        elif self.type == 4:
            self.pushButton_4.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                            "background-color: rgb(230, 230, 230);")
        elif self.type == 5:
            self.pushButton_5.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                            "background-color: rgb(230, 230, 230);")
        elif self.type == 6:
            self.pushButton_6.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                            "background-color: rgb(230, 230, 230);")
        elif self.type == 7:
            self.pushButton_7.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                            "background-color: rgb(230, 230, 230);")
        elif self.type == 8:
            self.pushButton_8.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                            "background-color: rgb(230, 230, 230);")
        elif self.type == 9:
            self.pushButton_9.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                            "background-color: rgb(230, 230, 230);")
        elif self.type == 10:
            self.pushButton_10.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                             "background-color: rgb(230, 230, 230);")
        else:
            pass
        self.type = 3

    # 버튼 4번 클릭시 그래프 생성
    def btnClick4(self):
        self.draw_graph(4)
        self.pushButton_4.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                        "background-color: rgb(255, 130, 130);")
        if self.type == 1:
            self.pushButton.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                          "background-color: rgb(230, 230, 230);")
        elif self.type == 2:
            self.pushButton_2.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                            "background-color: rgb(230, 230, 230);")
        elif self.type == 3:
            self.pushButton_3.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                            "background-color: rgb(230, 230, 230);")
        elif self.type == 5:
            self.pushButton_5.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                            "background-color: rgb(230, 230, 230);")
        elif self.type == 6:
            self.pushButton_6.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                            "background-color: rgb(230, 230, 230);")
        elif self.type == 7:
            self.pushButton_7.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                            "background-color: rgb(230, 230, 230);")
        elif self.type == 8:
            self.pushButton_8.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                            "background-color: rgb(230, 230, 230);")
        elif self.type == 9:
            self.pushButton_9.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                            "background-color: rgb(230, 230, 230);")
        elif self.type == 10:
            self.pushButton_10.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                             "background-color: rgb(230, 230, 230);")
        else:
            pass
        self.type = 4

    # 버튼 5번 클릭시 그래프 생성
    def btnClick5(self):
        self.draw_graph(5)
        self.pushButton_5.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                        "background-color: rgb(255, 130, 130);")
        if self.type == 1:
            self.pushButton.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                          "background-color: rgb(230, 230, 230);")
        elif self.type == 2:
            self.pushButton_2.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                            "background-color: rgb(230, 230, 230);")
        elif self.type == 3:
            self.pushButton_3.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                            "background-color: rgb(230, 230, 230);")
        elif self.type == 4:
            self.pushButton_4.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                            "background-color: rgb(230, 230, 230);")
        elif self.type == 6:
            self.pushButton_6.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                            "background-color: rgb(230, 230, 230);")
        elif self.type == 7:
            self.pushButton_7.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                            "background-color: rgb(230, 230, 230);")
        elif self.type == 8:
            self.pushButton_8.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                            "background-color: rgb(230, 230, 230);")
        elif self.type == 9:
            self.pushButton_9.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                            "background-color: rgb(230, 230, 230);")
        elif self.type == 10:
            self.pushButton_10.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                             "background-color: rgb(230, 230, 230);")
        else:
            pass
        self.type = 5

    # 버튼 6번 클릭시 그래프 생성
    def btnClick6(self):
        self.draw_graph(6)
        self.pushButton_6.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                        "background-color: rgb(255, 130, 130);")
        if self.type == 1:
            self.pushButton.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                          "background-color: rgb(230, 230, 230);")
        elif self.type == 2:
            self.pushButton_2.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                            "background-color: rgb(230, 230, 230);")
        elif self.type == 3:
            self.pushButton_3.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                            "background-color: rgb(230, 230, 230);")
        elif self.type == 4:
            self.pushButton_4.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                            "background-color: rgb(230, 230, 230);")
        elif self.type == 5:
            self.pushButton_5.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                            "background-color: rgb(230, 230, 230);")
        elif self.type == 7:
            self.pushButton_7.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                            "background-color: rgb(230, 230, 230);")
        elif self.type == 8:
            self.pushButton_8.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                            "background-color: rgb(230, 230, 230);")
        elif self.type == 9:
            self.pushButton_9.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                            "background-color: rgb(230, 230, 230);")
        elif self.type == 10:
            self.pushButton_10.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                             "background-color: rgb(230, 230, 230);")
        else:
            pass
        self.type = 6

    # 버튼 7번 클릭시 그래프 생성
    def btnClick7(self):
        self.draw_graph(7)
        self.pushButton_7.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                        "background-color: rgb(255, 130, 130);")
        if self.type == 1:
            self.pushButton.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                          "background-color: rgb(230, 230, 230);")
        elif self.type == 2:
            self.pushButton_2.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                            "background-color: rgb(230, 230, 230);")
        elif self.type == 3:
            self.pushButton_3.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                            "background-color: rgb(230, 230, 230);")
        elif self.type == 4:
            self.pushButton_4.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                            "background-color: rgb(230, 230, 230);")
        elif self.type == 5:
            self.pushButton_5.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                            "background-color: rgb(230, 230, 230);")
        elif self.type == 6:
            self.pushButton_6.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                            "background-color: rgb(230, 230, 230);")
        elif self.type == 8:
            self.pushButton_8.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                            "background-color: rgb(230, 230, 230);")
        elif self.type == 9:
            self.pushButton_9.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                            "background-color: rgb(230, 230, 230);")
        elif self.type == 10:
            self.pushButton_10.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                             "background-color: rgb(230, 230, 230);")
        else:
            pass
        self.type = 7

    # 버튼 8번 클릭시 그래프 생성
    def btnClick8(self):
        self.draw_graph(8)
        self.pushButton_8.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                        "background-color: rgb(255, 130, 130);")
        if self.type == 1:
            self.pushButton.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                          "background-color: rgb(230, 230, 230);")
        elif self.type == 2:
            self.pushButton_2.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                            "background-color: rgb(230, 230, 230);")
        elif self.type == 3:
            self.pushButton_3.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                            "background-color: rgb(230, 230, 230);")
        elif self.type == 4:
            self.pushButton_4.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                            "background-color: rgb(230, 230, 230);")
        elif self.type == 5:
            self.pushButton_5.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                            "background-color: rgb(230, 230, 230);")
        elif self.type == 6:
            self.pushButton_6.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                            "background-color: rgb(230, 230, 230);")
        elif self.type == 7:
            self.pushButton_7.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                            "background-color: rgb(230, 230, 230);")
        elif self.type == 9:
            self.pushButton_9.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                            "background-color: rgb(230, 230, 230);")
        elif self.type == 10:
            self.pushButton_10.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                             "background-color: rgb(230, 230, 230);")
        else:
            pass
        self.type = 8

    # 버튼 9번 클릭시 그래프 생성
    def btnClick9(self):
        self.draw_graph(9)
        self.pushButton_9.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                        "background-color: rgb(255, 130, 130);")
        if self.type == 1:
            self.pushButton.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                          "background-color: rgb(230, 230, 230);")
        elif self.type == 2:
            self.pushButton_2.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                            "background-color: rgb(230, 230, 230);")
        elif self.type == 3:
            self.pushButton_3.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                            "background-color: rgb(230, 230, 230);")
        elif self.type == 4:
            self.pushButton_4.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                            "background-color: rgb(230, 230, 230);")
        elif self.type == 5:
            self.pushButton_5.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                            "background-color: rgb(230, 230, 230);")
        elif self.type == 6:
            self.pushButton_6.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                            "background-color: rgb(230, 230, 230);")
        elif self.type == 7:
            self.pushButton_7.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                            "background-color: rgb(230, 230, 230);")
        elif self.type == 8:
            self.pushButton_8.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                            "background-color: rgb(230, 230, 230);")
        elif self.type == 10:
            self.pushButton_10.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                             "background-color: rgb(230, 230, 230);")
        else:
            pass
        self.type = 9

    # 버튼 10번 클릭시 그래프 생성
    def btnClick10(self):
        self.draw_graph(10)
        self.pushButton_10.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                         "background-color: rgb(255, 130, 130);")
        if self.type == 1:
            self.pushButton.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                          "background-color: rgb(230, 230, 230);")
        elif self.type == 2:
            self.pushButton_2.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                            "background-color: rgb(230, 230, 230);")
        elif self.type == 3:
            self.pushButton_3.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                            "background-color: rgb(230, 230, 230);")
        elif self.type == 4:
            self.pushButton_4.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                            "background-color: rgb(230, 230, 230);")
        elif self.type == 5:
            self.pushButton_5.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                            "background-color: rgb(230, 230, 230);")
        elif self.type == 6:
            self.pushButton_6.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                            "background-color: rgb(230, 230, 230);")
        elif self.type == 7:
            self.pushButton_7.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                            "background-color: rgb(230, 230, 230);")
        elif self.type == 8:
            self.pushButton_8.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                            "background-color: rgb(230, 230, 230);")
        elif self.type == 9:
            self.pushButton_9.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                            "background-color: rgb(230, 230, 230);")
        else:
            pass
        self.type = 10

    #그래프저장 함수
    def save_graph(self):
        reply = QMessageBox.question(self, 'Glaph Save', '그래프를 저장하시겠습니까? 저장되는 그래프 파일명은 stock_graph.png이며,'
                                                         ' 저장되는 위치는 현재 Pycharm 폴더입니다.',
                                     QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.Yes:
            plt.figure(self.fig.number)
            plt.savefig('stock_graph.png')
        else:
            return

    #github 연결함수(data 제공)
    def connect_git(self):
        reply = QMessageBox.question(self, 'Data Update', 'Github 사이트에 연결하시겠습니까?',
                                     QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.Yes:
            url = 'https://github.com/ashildlee/project'
            webbrowser.open(url)
        else:
            return

    #그래프 그리는함수
    def draw_graph(self, index):
        plt.clf()
        mpl.rcParams['font.family'] = 'Pyunji R'
        ax = self.fig.add_subplot(1, 1, 1)  # subplot 생성
        ax.plot(self.df_search_extended['일자'], self.df_search_extended[self.result_arr[index - 1]['code']],
                label=self.result_arr[index - 1]['name'],
                color='b')
        ax.yaxis.set_major_formatter(mpl.ticker.StrMethodFormatter('{x:,.0f}'))
        ax.set_xlabel('Date', fontsize=8)  # x축 설정
        ax.set_ylabel('Stock Price', fontsize=6)  # y축 설정
        ax.grid()
        plt.xticks(rotation=45, fontsize=8)
        plt.yticks(rotation=45, fontsize=8)
        ax1 = ax.twinx()
        ax1.plot(self.dfkospi_extended['일자'], self.dfkospi_extended['kospi'], label='Kospi Index', color='r')
        ax1.yaxis.set_major_formatter(mpl.ticker.StrMethodFormatter('{x:,.0f}'))
        ax.set_title(self.result_arr[index - 1]['name']+"'s Stock Price & KOSPI Index ", fontsize=16)  # 타이틀 설정
        ax.set_xlabel('Date', fontsize=12)  # x축 설정
        ax.set_ylabel('Stock Price', fontsize=6)  # y1축 설정
        ax1.set_ylabel('Kospi Index', fontsize=6)  # y2축 설정
        ax.legend(loc='upper left', fontsize=10)
        ax1.legend(loc='upper right', fontsize=10)
        plt.yticks(rotation=45, fontsize=8)
        plt.axvspan(self.searched_date_start, self.searched_date_end, facecolor='salmon', alpha=0.2)
        self.canvas.draw()


    # 매수/매도 함수
    def radioBuysellFunction(self):
        if self.radio_buy.isChecked():
            print("Radio buy Checked")
        elif self.radio_sell.isChecked():
            print("Radio sell Checked")

    # 검색 함수
    def btnsearchFunction(self):
        self.pushButton.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                      "background-color: rgb(230, 230, 230);")
        self.pushButton_2.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                        "background-color: rgb(230, 230, 230);")
        self.pushButton_3.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                        "background-color: rgb(230, 230, 230);")
        self.pushButton_4.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                        "background-color: rgb(230, 230, 230);")
        self.pushButton_5.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                        "background-color: rgb(230, 230, 230);")
        self.pushButton_6.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                        "background-color: rgb(230, 230, 230);")
        self.pushButton_7.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                        "background-color: rgb(230, 230, 230);")
        self.pushButton_8.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                        "background-color: rgb(230, 230, 230);")
        self.pushButton_9.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                        "background-color: rgb(230, 230, 230);")
        self.pushButton_10.setStyleSheet("font: 10pt \"휴먼편지체\";\n"
                                         "background-color: rgb(230, 230, 230);")


        try:
            if self.date_start.date() > self.date_end.date():
                print("error")
                QMessageBox.critical(self, "Error", '시작날짜가 종료날짜보다 미래날짜로 잘못 지정 되었습니다. 다시 선택하세요')
                return

            self.result_arr = self.data_manager.search(self.date_start.date().toString("yyyy-MM-dd"),
                                                  self.date_end.date().toString("yyyy-MM-dd"),
                                                  self.getRadioSupplier(),
                                                  self.getRadioBuysell())

            self.tableWidget.setRowCount(0) #Gui상 표위젯 만들기
            for idx, item in enumerate(self.result_arr):
                self.tableWidget.insertRow(idx)
                self.tableWidget.setItem(idx, 0, QTableWidgetItem(item["code"]))
                self.tableWidget.setItem(idx, 1, QTableWidgetItem(item["name"]))
                self.tableWidget.setItem(idx, 2, QTableWidgetItem(item["date_count"]))
                self.tableWidget.setItem(idx, 3, QTableWidgetItem(item["rate"]))


            self.searched_date_start = self.date_start.date().toString("yyyy-MM-dd")
            self.searched_date_end = self.date_end.date().toString("yyyy-MM-dd")

            mask = (self.dfClose['일자'] >= (
                        datetime.strptime(self.date_start.date().toString("yyyy-MM-dd"), "%Y-%m-%d") - relativedelta(
                    months=1)).strftime("%Y-%m-%d")) & (
                           self.dfClose['일자'] <= (
                               datetime.strptime(self.date_end.date().toString("yyyy-MM-dd"), "%Y-%m-%d") + relativedelta(
                           months=1)).strftime("%Y-%m-%d"))
            self.df_search_extended = self.dfClose.loc[mask]
            self.dfkospi_extended = self.dfkospi.loc[mask]

            self.pushButton.show()
            self.pushButton_2.show()
            self.pushButton_3.show()
            self.pushButton_4.show()
            self.pushButton_5.show()
            self.pushButton_6.show()
            self.pushButton_7.show()
            self.pushButton_8.show()
            self.pushButton_9.show()
            self.pushButton_10.show()
            self.pushButton_11.show()

        except:
            print("exception")
            QMessageBox.critical(self, "Error", '해당 기간의 주식 Data가 없습니다. 날짜 지정을 다시 선택하세요')

    # 라디오 함수
    def radioSupplierFunction(self):
        if self.radio_supplier_1.isChecked():
            print("Radio 1 Checked")
        elif self.radio_supplier_2.isChecked():
            print("Radio 2 Checked")
        elif self.radio_supplier_3.isChecked():
            print("Radio 3 Checked")

    # 날짜 함수 - 시작
    def dateStartFunction(self):
        print(self.date_start.date())

    # 날짜 함수 - 종료
    def dateEndFunction(self):
        print(self.date_end.date())

    def selectTableFunction(self):
        print(self.tableWidget.currentRow())
        select_row = self.tableWidget.currentRow()
        print("selected stockname : ", self.tableWidget.item(select_row, 0).text())

    # 매수/매도 값 가져오기
    def getRadioBuysell(self):
        if self.radio_buy.isChecked():
            return 1
        elif self.radio_sell.isChecked():
            return 2

    # 라디오 값 가져오기
    def getRadioSupplier(self):
        if self.radio_supplier_1.isChecked():
            return 1
        elif self.radio_supplier_2.isChecked():
            return 2
        elif self.radio_supplier_3.isChecked():
            return 3

    def closeEvent(self, event):
        reply = QMessageBox.question(self, 'Exit the program', '프로그램을 종료 하시겠습니까?',
                                     QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.Yes:
            event.accept()
        else:
            event.ignore()

if __name__ == "__main__":
    app = QApplication(sys.argv)
    myWindow = WindowClass()
    myWindow.show()
    app.exec_()